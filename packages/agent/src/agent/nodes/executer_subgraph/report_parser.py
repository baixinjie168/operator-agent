"""Parse ATK execution artifacts into :class:`ExecutionResult` data.

Three responsibilities:

1. Download the ``report/`` directory + ``log/atk.log`` from the remote
   machine via SFTP.
2. Parse the latest ``.xlsx`` in ``report/`` row-by-row, mapping the
   well-known columns (``运行结果`` / ``失败原因`` / ``用例JSON信息`` …)
   onto :class:`ReportRecord` fields.
3. Aggregate everything into a :class:`TaskReportData` ready to embed
   in :class:`ExecutionResult`.

All functions are tolerant: parsing errors are captured into
``parse_error`` / ``error_message`` fields and never raised, so that
result-extraction failures don't terminate the main execution flow.
"""

from __future__ import annotations

import asyncio
import json
import logging
import re
from pathlib import Path
from typing import Any

import asyncssh
from openpyxl import load_workbook

from .execution_result import ReportRecord, TaskReportData

logger = logging.getLogger(__name__)


# ── Column matching ─────────────────────────────────────────────────────────
#
# ATK reports are produced by an internal tool whose column titles drift
# between versions.  We fuzzy-match against these canonical names so the
# parser keeps working across minor report-schema changes.

_RUN_RESULT_ALIASES = (
    "运行结果", "运行状态", "result", "run_result", "status", "Result",
    "测试结果", "结果",
)
_FAILURE_REASON_ALIASES = (
    "失败原因", "原因", "error", "fail_reason", "failure_reason",
    "失败信息", "error_message",
)
_CASE_JSON_ALIASES = (
    "用例JSON信息", "用例 JSON 信息", "用例json信息", "case_json",
    "用例JSON", "case_info", "用例信息", "case", "json",
)
_ID_ALIASES = (
    "用例ID", "用例 ID", "id", "ID", "case_id", "用例编号", "序号",
)


def _norm(value: Any) -> str:
    """Lowercased, whitespace-collapsed string for fuzzy column matching."""
    return re.sub(r"\s+", "", str(value or "").strip().lower())


def _match_column(header: str, aliases: tuple[str, ...]) -> bool:
    target = _norm(header)
    for alias in aliases:
        if _norm(alias) == target:
            return True
    # Substring fallback for column names like "运行结果 (pass/fail)"
    return any(_norm(a) in target for a in aliases if len(_norm(a)) >= 2)


def _truthy_pass(value: Any) -> bool:
    """Interpret the run_result cell as a pass / fail boolean."""
    if value is None:
        return False
    text = _norm(value)
    if not text:
        return False
    if text in {"pass", "passed", "success", "ok", "成功", "通过", "1", "true", "yes", "y"}:
        return True
    if text in {"fail", "failed", "error", "fail_case", "失败", "未通过", "0", "false", "no", "n"}:
        return False
    # Anything containing 成功 / 通过 → pass; 失败 → fail; else unknown = fail
    if "失败" in text or "fail" in text or "error" in text:
        return False
    if "成功" in text or "通过" in text or "pass" in text:
        return True
    return False


# ── Local cache directory ───────────────────────────────────────────────────

def make_local_cache_dir(project_root: Path, operator_name: str, run_id: str) -> Path:
    """Create a per-run directory for downloaded ATK artifacts.

    Layout: ``<project_root>/execution_results/<operator_name>/<run_id>/``
    so each operator keeps its own folder and multiple runs of the same
    operator do not overwrite each other.
    """
    safe_operator = re.sub(r"[^A-Za-z0-9_.-]", "_", operator_name).strip("_") or "operator"
    safe_run = re.sub(r"[^A-Za-z0-9_.-]", "_", run_id)[:48] or "run"
    cache = project_root / "execution_results" / safe_operator / safe_run
    cache.mkdir(parents=True, exist_ok=True)
    return cache


# ── Remote download ─────────────────────────────────────────────────────────

async def sftp_download_file(
    conn: asyncssh.SSHClientConnection,
    remote_path: str,
    local_path: Path,
) -> None:
    """Pull a single remote file via SFTP. Swallows missing-file errors."""
    try:
        async with conn.start_sftp_client() as sftp:
            await sftp.get(remote_path, str(local_path))
    except FileNotFoundError:
        logger.warning("report_parser: remote file not found: %s", remote_path)
    except Exception as e:
        logger.warning("report_parser: sftp get %s failed: %s", remote_path, e)


async def sftp_list_dir(conn: asyncssh.SSHClientConnection, remote_dir: str) -> list[str]:
    """List entries in a remote directory. Returns [] on any failure."""
    try:
        async with conn.start_sftp_client() as sftp:
            entries = await sftp.listdir(remote_dir)
        return [str(e) for e in entries]
    except Exception as e:
        logger.warning("report_parser: listdir %s failed: %s", remote_dir, e)
        return []


# ── xlsx parsing ────────────────────────────────────────────────────────────

def _find_latest_xlsx(report_dir: Path) -> Path | None:
    """Return the newest ``.xlsx`` in ``report_dir`` by mtime, or None."""
    if not report_dir.is_dir():
        return None
    candidates = sorted(
        report_dir.glob("*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def parse_xlsx_report(report_dir: Path) -> TaskReportData:
    """Parse the latest xlsx in ``report_dir`` into a :class:`TaskReportData`.

    Accepts either a directory (picks the newest ``*.xlsx``) or a single
    xlsx file path (parses that file directly).

    On any error, returns a :class:`TaskReportData` with ``parse_error``
    set instead of raising — extraction failures must not abort the
    main execution flow.
    """
    data = TaskReportData()

    if report_dir.is_file() and report_dir.suffix.lower() == ".xlsx":
        latest = report_dir
    else:
        latest = _find_latest_xlsx(report_dir)

    if latest is None:
        data.parse_error = "report 目录下未找到任何 .xlsx 文件"
        return data
    data.report_path = str(latest)

    try:
        wb = load_workbook(latest, read_only=True, data_only=True)
    except Exception as e:
        data.parse_error = f"无法打开 xlsx: {e}"
        return data

    try:
        ws = wb.active
        data.sheet_name = ws.title

        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            data.parse_error = "xlsx 为空, 无表头"
            return data

        # Map: canonical field → column index
        col_for: dict[str, int] = {}
        for idx, header in enumerate(header_row):
            if _match_column(header or "", _ID_ALIASES):
                col_for.setdefault("id", idx)
            elif _match_column(header or "", _RUN_RESULT_ALIASES):
                col_for.setdefault("run_result", idx)
            elif _match_column(header or "", _FAILURE_REASON_ALIASES):
                col_for.setdefault("failure_reason", idx)
            elif _match_column(header or "", _CASE_JSON_ALIASES):
                col_for.setdefault("case_json", idx)

        # Iterate body rows
        for row in rows:
            if row is None:
                continue
            # Skip fully empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            def _cell(key: str) -> Any:
                idx = col_for.get(key)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            case_json_raw = _cell("case_json")
            case_json_obj: dict[str, Any] | None = None
            if case_json_raw not in (None, ""):
                if isinstance(case_json_raw, (dict, list)):
                    case_json_obj = case_json_raw if isinstance(case_json_raw, dict) else {"items": case_json_raw}
                else:
                    try:
                        parsed = json.loads(str(case_json_raw))
                        if isinstance(parsed, dict):
                            case_json_obj = parsed
                        elif isinstance(parsed, list):
                            case_json_obj = {"items": parsed}
                        else:
                            case_json_obj = {"value": parsed}
                    except (json.JSONDecodeError, TypeError, ValueError):
                        case_json_obj = {"raw": str(case_json_raw)}

            # Prefer the test case's DB id (carried inside case_json.id),
            # since that's the authoritative id used elsewhere in the
            # pipeline (exec_results.case_id, test_cases.id, etc.).  Fall
            # back to the xlsx's id-style column when case_json has no id
            # — some ATK reports omit it.
            xlsx_id = (str(_cell("id")).strip() if _cell("id") is not None else "") or ""
            case_json_id = ""
            if isinstance(case_json_obj, dict):
                raw_cj_id = case_json_obj.get("id")
                if raw_cj_id is not None and str(raw_cj_id).strip():
                    case_json_id = str(raw_cj_id).strip()
            record_id = case_json_id or xlsx_id or None

            record = ReportRecord(
                id=record_id,
                run_result=(str(_cell("run_result")).strip() if _cell("run_result") is not None else None),
                failure_reason=(str(_cell("failure_reason")).strip() if _cell("failure_reason") is not None else None),
                case_json=case_json_obj,
                extra={
                    str(header_row[i] or f"col_{i}"): (row[i] if i < len(row) else None)
                    for i in range(len(header_row))
                    if i not in col_for.values()
                },
            )

            data.report_records.append(record)
            if _truthy_pass(record.run_result):
                data.passed += 1
            else:
                data.failed += 1
        data.record_count = len(data.report_records)
    except Exception as e:
        logger.exception("report_parser: xlsx parsing failed for %s", latest)
        data.parse_error = f"xlsx 解析失败: {e}"
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return data


# ── Convenience: pull everything ────────────────────────────────────────────

async def collect_remote_artifacts(
    conn: asyncssh.SSHClientConnection,
    remote_output_dir: str,
    cache_dir: Path,
) -> tuple[TaskReportData, str, Path | None]:
    """Download report + log artifacts and parse the xlsx.

    The remote layout still groups xlsx files under ``report/`` (that's
    ATK's canonical structure and we can't change it), but locally we
    flatten everything into ``cache_dir`` so the operator's per-run
    folder holds ``atk.log``, ``result.json`` and the xlsx reports
    side-by-side.

    Returns ``(report_data, log_content, latest_xlsx_local_path)``.
    Never raises; errors are captured into the returned object.
    """
    remote_report_dir = f"{remote_output_dir}/report"
    remote_log_path = f"{remote_output_dir}/log/atk.log"

    local_log_path = cache_dir / "atk.log"
    cache_dir.mkdir(parents=True, exist_ok=True)

    # Best-effort: list entries and pull each one straight into cache_dir,
    # preserving ATK's original file names.
    remote_entries = await sftp_list_dir(conn, remote_report_dir)
    for entry in remote_entries:
        await sftp_download_file(conn, f"{remote_report_dir}/{entry}", cache_dir / entry)

    await sftp_download_file(conn, remote_log_path, local_log_path)

    # Parse xlsx — ``parse_xlsx_report`` globs ``*.xlsx`` in whatever
    # directory we hand it, so passing cache_dir directly is equivalent
    # to passing the old ``cache_dir/report`` subdir.
    report_data = parse_xlsx_report(cache_dir)
    latest_xlsx = (
        Path(report_data.report_path) if report_data.report_path else None
    )

    log_content = ""
    if local_log_path.exists():
        try:
            log_content = local_log_path.read_text(encoding="utf-8", errors="replace")
            if len(log_content) > 200_000:
                # Truncate to keep SSE / DB payloads manageable
                log_content = log_content[:200_000] + "\n... [log truncated]"
        except Exception as e:
            logger.warning("report_parser: failed to read atk.log: %s", e)

    return report_data, log_content, latest_xlsx


__all__ = [
    "collect_remote_artifacts",
    "make_local_cache_dir",
    "parse_xlsx_report",
    "sftp_download_file",
    "sftp_list_dir",
]